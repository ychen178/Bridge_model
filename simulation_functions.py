#------------ import 
from multiprocessing import reduction
import os
import numpy as np
import matplotlib.pyplot as plt
import time
import math
import scipy
from scipy import stats, spatial
from scipy.linalg import qr
from scipy.special import logit, expit
import pandas as pd
from operator import add, length_hint 
import itertools
import sklearn
from sklearn import preprocessing
from sklearn.model_selection import KFold
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.impute import SimpleImputer
from sklearn.metrics import confusion_matrix, mean_squared_error,precision_recall_fscore_support
from sklearn.preprocessing import OrdinalEncoder, OneHotEncoder
from sklearn.decomposition import PCA
from sklearn.manifold import TSNE
import statsmodels.api as sm

import copy
import inspect
import pickle
from enum import Enum
import seaborn as sns
import umap
import random

import torch 
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim 
from torch.utils.data import Dataset, DataLoader, random_split, Subset
import torch.nn.utils.parametrizations as parametrizations

from typing import Union
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo 
import csv
import datetime
import heapq


from bridge_functions import *


#------------ Simulation functions


## Simulation function with Z from Normal, can simulate X as continuous variable or indicator variables.
# since we care about the loss for all Ys not just the ones observed in panels. Hence Y will be full data but we will use mask objects to create the training data --- "mask_ind" to mask unobserved genes in each panel (not including mask1 entries if mask1_percent > 0), "mask1" will be among observed genes randomly sample by "mask1_percent" for early stopping, "mask2" will be the full Y to evaluate (what we care about)
# px: number of centers
# gamma_cont: parameter for the continous covariate, gamma_cont: dimension of n_gene * n_Vcont
# gamma_binary: parameter for the binary covariate, gamma_cont: dimension of n_gene * n_Vbinary
# nV_cat: number of categorical covariates
# center_prop: a vector, proportion of each center, and proportion of NA gene columns in each center
# gene_prop: a list of arrays with length px, each entry is the coverage of genes for this center, e.g., (0.2, 0.5) is that this center observed genes from gene 20 to gene 50 if there are 100 genes
# will generate the mask_obj for train_model() function
# "seed": will be used to simulate the data
# evaluate_on: "all" for all observed+unobserved genes, "overlap":overlapping genes. Will use this to create mask2_ind in mask_obj accordingly. So don't need mask2_percent or mask2_seed
## output:
# mask_obj: use it to train the propsoed model
# mask_obj_overlap: use it to train the model with overlapping genes only
# mask_obj_full: use it to train the model with full observations of all genes
# O_coef: coefficient of Z on outcome 
# O_coef_V: coefficient of V on outcome

def sim_wCovariates(n_sub, center_prop, beta, W, gamma_cont=None, gamma_binary=None, nV_cont=1, nV_binary=1, Z_sd=1, V_sd=1, gene_prop=None, seed=1, evaluate_on="all", mask1_seed=1, mask1_percent=0, 
O_coef=None, O_coef_V=None, O_scale=1, WZ_form="linear", Z_dist="normal"):
    
    # set seed    
    np.random.seed(seed)

    n_gene = W.shape[0]
    n_factor = W.shape[1]
    px = len(center_prop)

    if Z_dist == "normal":
        # simulate Z from Normal
        Z = np.random.normal(loc = 0, scale=Z_sd, size = (n_sub, n_factor))
    elif Z_dist == "uniform":
        Z = np.random.uniform(low=-20, high=20, size = (n_sub, n_factor))
    
    # Create X as indicators for each center
    X = np.zeros(shape = (n_sub, px))
    X_comb = np.empty(shape = (n_sub))
    start = 0
    for j in range(px):
        end = math.floor(start + n_sub * center_prop[j])
        X[start: end, j] = 1
        X_comb[start:end] = j 
        start = end
    
    # simulate covariates V and its coeficients (always simualte V, but not to include in the latent model if gamma coefs are None)
    V_cont = np.random.normal(loc = 0, scale=V_sd, size = (n_sub, nV_cont))
    V_binary = np.random.binomial(1, 0.5, size = (n_sub, nV_binary))
    V = np.concatenate((V_cont, V_binary), axis=1)


    V_terms_sum = 0
    if gamma_cont is not None:
        V_term_cont = V_cont @ gamma_cont.T
        V_terms_sum = V_terms_sum + V_term_cont
    else:
        V_term_cont = None
    if gamma_binary is not None:
        V_term_binary = V_binary @ gamma_binary.T
        V_terms_sum = V_terms_sum + V_term_binary
    else:
        V_term_binary = None

    # combine to one V matrix
    if gamma_cont is not None and gamma_binary is not None:
        gamma = np.concatenate((gamma_cont, gamma_binary), axis=1)
    elif gamma_cont is not None:
        gamma = gamma_cont
    elif gamma_binary is not None:
        gamma = gamma_binary
    else:
        gamma = None



    # calculate the linear 
    if WZ_form == "linear":
        WZ = Z @ W.T
    if WZ_form == "quadratic":
        WZ = Z @ W.T + np.square(Z) @ W.T / 10  # since the scale is too big after square compared to xb
        # WZ = np.square(Z) @ W.T / 20  # old: too bad for out of sample on test


    Xbeta = X @ beta.T
    linear = Xbeta + WZ + V_terms_sum
    prob = 1 / (1 + np.exp(-linear))
    prob_obs = copy.copy(prob)
    # corrected under the proposed model 
    beta_mean = beta.mean(1)
    linear_correct = beta_mean + Z @ W.T
    prob_correct = pd.DataFrame(1 / (1 + np.exp(-linear_correct)))
    prob_correct_binary = (prob_correct > 0.5) * 1

    
    # simulate binary outcome based on prob
    # Will make Y_obs, prob_obs to NA for unobserved gene entries.
    Y = pd.DataFrame(np.random.binomial(1, prob).astype('float'))
    Y_obs = copy.copy(Y)
    Y_overlap = copy.copy(Y)

    # sum of Yprob or Y by center (for all genes, non-corrected & corrected)
    Ysum_byX = Y.groupby(X_comb).mean().sum(1).values  # mean across subjects then sum over genes
    YprobSum_byX = pd.DataFrame(prob).groupby(X_comb).mean().sum(1).values  # mean across subjects then sum over genes
    # corrected
    YprobSum_byX_c = prob_correct.groupby(X_comb).mean().sum(1).values  # mean across subjects then sum over genes
    Ysum_byX_c = prob_correct_binary.groupby(X_comb).mean().sum(1).values  # mean across subjects then sum over genes


    ## Create the mask objects manually to be used in train_model() -- 3 objects for 3 sets of models
    # mask_ind==0 entries will be used for training the model (mask_ind==1 include unobserved entries + mask1_ind==1 entries). 
    # mask1_ind==1 -- randomly pick from observed gene entries (will evaluate loss for these mask1_ind==1 entries for early stopping)
    # mask2_ind==1 entries will be used to evaluate loss_mask1. mask2_ind=1 for every entry in Y if evaluate_on == "all" (both observed and unobserved) will evaluate on all entries)

    # first use the create_mask_Y() function to initiate the mask_obj (empty object), will create mask indicators manually later
    mask_obj = create_mask_Y(Y, mask1_seed = mask1_seed, mask1_percent = mask1_percent)
    mask_obj.mask_ind = pd.DataFrame(np.full_like(Y, 0))
    # "mask_obj_overlap" for using only the overlapping genes
    mask_obj_overlap = create_mask_Y(Y, mask1_seed = mask1_seed, mask1_percent = mask1_percent)
    mask_obj_overlap.mask_ind = pd.DataFrame(np.full_like(Y, 0))
    # for using full data in training the model
    mask_obj_full = create_mask_Y(Y, mask1_seed = mask1_seed, mask1_percent = mask1_percent)
    mask_obj_full.mask_ind = pd.DataFrame(np.full_like(Y, 0))

    # create mask_ind based on "gene_prop" + make Y_obs with NA entries (used in model training)
    if gene_prop is not None:

        gene_panel_ind = np.zeros((px, n_gene))
        beta_real = copy.deepcopy(beta)  # make those unobserved entries to NA -- beta_real will be the informative beta values then

        for i in range(px):
            col_idx_start_end = np.round(n_gene * gene_prop[i]).astype('int')
            col_idx = np.array(range(col_idx_start_end[0], col_idx_start_end[1]))
            na_col_idx = np.array(list(set(range(n_gene)) - set(col_idx)))
            if len(na_col_idx) > 0:
                Y_obs.iloc[np.where(X[:,i]==1)[0], na_col_idx] = np.nan  
                prob_obs[np.where(X[:,i]==1)[0].reshape(-1,1), na_col_idx] = np.nan # # Note the way of writting row and col selection for np.array
                beta_real[na_col_idx, i] = np.nan
                mask_obj.mask_ind.iloc[np.where(X[:,i]==1)[0], na_col_idx] = 1
                Y_overlap.iloc[:, na_col_idx] = np.nan
                mask_obj_overlap.mask_ind.iloc[:, na_col_idx] = 1
            gene_panel_ind[i, col_idx] = 1  # each row is for one panel, 1 for the genes covered in this panel
            # mask_obj.mask_ind[np.where(X[:,i]==1)[0].reshape(-1,1), na_col_idx] = 1 # if mask_ind is dataframe
    else:
        gene_panel_ind = None

    if evaluate_on=="all":
        # all entries in Y will be mask2
        mask_obj.mask2_ind = pd.DataFrame(np.full_like(Y, 1))
        mask_obj_overlap.mask2_ind = pd.DataFrame(np.full_like(Y, 1))
        mask_obj_full.mask2_ind = pd.DataFrame(np.full_like(Y, 1))

        ## if mask1_percent > 0, create "mask1_ind" among the observed entries (entries that mask_ind==0), and then make those entries to have mask_ind==1 (so not to be used in model training)
        if mask1_percent > 0:
            # for "mask_obj" create mask1_ind 
            nonmask1_entry = mask_obj.mask_ind == 1 # mask among the observed genes (mask_ind==0)
            mask_obj.mask1_ind = mask_obj._create_mask_ind(mask_percent=mask1_percent, nonmask_entry = nonmask1_entry, seed = mask1_seed)
            mask_obj.mask_ind[mask_obj.mask1_ind==1] = 1 # make mask1 entries also mask_ind==1
            # for "mask_obj_overlap"
            nonmask1_entry = mask_obj_overlap.mask_ind == 1
            mask_obj_overlap.mask1_ind = mask_obj_overlap._create_mask_ind(mask_percent=mask1_percent, nonmask_entry = nonmask1_entry, seed = mask1_seed)
            mask_obj_overlap.mask_ind[mask_obj_overlap.mask1_ind==1] = 1 
            # for "mask_obj_full"
            mask_obj_full.mask1_ind = mask_obj_overlap._create_mask_ind(mask_percent=mask1_percent, seed = mask1_seed)
            mask_obj_full.mask_ind[mask_obj_full.mask1_ind==1] = 1


    # loss (for the observed) under the true parameter
    loss_full = -np.nanmean(Y_obs * np.log(prob) + (1 - Y_obs) * np.log(1-prob))
    # accuracy under the true parameter
    accuracy_full = np.nanmean((Y_obs.values == (prob > 0.5))[~np.isnan(Y_obs)])  # Note: need to filter out NA first since na==1 is 0


    # make Y and Y_obs to df
    Y = pd.DataFrame(Y)
    Y_obs = pd.DataFrame(Y_obs)
    Y_overlap = pd.DataFrame(Y_overlap)


    # create outcome O as function as Z & X, O_coef is of size = (n_factor, 1) 
    if O_coef is None and O_coef_V is None:
        O = None
    elif  O_coef is None:
        O = np.random.normal(loc = V @ O_coef_V, scale=O_scale)
    elif O_coef_V is None:
        O = np.random.normal(loc = Z @ O_coef, scale=O_scale)
    else:
        O = np.random.normal(loc = Z @ O_coef + V @ O_coef_V, scale=O_scale)


    # # make X to dataframe
    X = pd.DataFrame(X, columns=['MSK', 'DFCI'])  # so that the check_beta_concordance will work
    # X_comb = pd.Series(X_comb, name='institution')  # need to be pandas series


    # make those non-informative beta to NA
    real_beta = copy.deepcopy(beta)
    beta_update = np.empty_like(beta)
    for k in range(X.shape[1]):
        beta_update[:, k] = 1 - np.isnan(np.nanmean(Y_obs.loc[X.iloc[:, k]==1, ], 0))
    real_beta[beta_update==0] = np.nan



    return {'Z':Z, 'X':X, 'X_comb':X_comb, 'V':V,
            'Yprob':prob, 'Yprob_obs': prob_obs, 
            'Y':Y, 'Y_obs':Y_obs, 'Y_overlap':Y_overlap,
            'loss_full':loss_full, 'accuracy_full':accuracy_full, # for the observed entries
            'gene_panel_ind': gene_panel_ind,
            'mask_obj': mask_obj, 'mask_obj_overlap':mask_obj_overlap, 'mask_obj_full':mask_obj_full,
            'n_factor':n_factor, 'n_sub':n_sub, 'n_gene':n_gene, 'px':px,
            'beta':real_beta, 'beta_real':beta_real, 'W':W, 'gamma':gamma, 'Z_sd':Z_sd,
            'center_prop':center_prop, 'gene_prop':gene_prop, 
            'seed': seed, 'evaluate_on':evaluate_on,
            'mask1_seed': mask1_seed, 'mask1_percent':mask1_percent,
            'Ysum_byX':Ysum_byX, 'YprobSum_byX':YprobSum_byX, 
            'YprobSum_byX_c':YprobSum_byX_c, 'Ysum_byX_c':Ysum_byX_c,
            'Yprob_c':prob_correct, 'Yprob_c_binary':prob_correct_binary, 
            'beta_mean': beta_mean, 
            'WZ': WZ,  'Xbeta':Xbeta, 'V_term':V_terms_sum, 'V_term_cont':V_term_cont, 'V_term_binary':V_term_binary, 
            'O_coef':O_coef, 'O_coef_V':O_coef_V,  'O':O, 'WZ_form':WZ_form
            }



## create a excel file to store model results (simulation)
def create_modelres_excel_s(file, new_file=True, excel_tab="Sheet"):

    if new_file:
        wb = Workbook()  # create new excel file
    else:  # read in the existing excel file
        workbook_name = file
        wb = load_workbook(workbook_name)
    # create a new sheet in the excel 
    wb.create_sheet(excel_tab)
    page = wb[excel_tab]


    # add column headings.
    page.append(["Date", "Time", "model_type", "Seed", "WZ_form",
    "mask2_for", "optimal_loss_full",
    "loss0_train", "loss_train", 
    "loss0_mask1", "loss_mask1",
    "loss0_full", "loss_full",
    # minimal loss on mask1, the epoch, how many times of increase
    "min_loss_mask1", "min_loss0_mask1_epoch", "n_increase_bf_min_loss0_mask1"] + 
    ["ZWloss_train", "ZWloss_mask1", "ZWloss_mask2",
    "beta_loss", "gamma1_loss", "gamma2_loss",
    'min_beta_loss', 'min_beta_loss_epoch',
    'min_gamma_loss', 'min_gamma_loss_epoch',
    'min_ZW_loss0_mask1', 'min_ZW_loss0_mask1_epoch'
    ] +
    ["beta_sign_same", "beta_center_concordance", "beta_rmse"] + 
    ["var_explained_trueZ", "var_explained_estZ", "var_explained_V","var_explained_VY"] + 
    ["rmse_trueZ", "rmse_estZ", "rmse_V","rmse_VY"] + 
    ["Z_orthogonal", "min_delta",
    "dat_seed", "mask1_seed", "mask1_percent",
    "n_gene", "n_factor", "px", "pv", 
    "W_orthonormal", "beta_W_orth", "gamma_W_orth",
    "Z_demean_bycenter", "Z_scale_bycenter", 
    "penalty", "nquantile", "",
    "l1_gamma", "l1_W", "l1_beta",
    "weight_Y",
    "lr", "n_epochs", "early_stop", "patience",
    ] + 
    ["", "n_sub", "px", "Z_sd", "center_prop", "", "gene_prop"]
    )
    wb.save(file)
    return wb


## function to append model results to existing excel file (for simulation)
# preYprob_c, preY_c: "predicted_Yprob_corrected", the dataframe of mean for each gene (including tmb) by center
# obs_Y: used to identify the observed entries -- to identify real_beta -- will be different for different model (unobserved entries are NA)
# "model_type": is a string, will be output to the result excel file
def append_modelres_excel_s(model, model_train, mask_obj, model_type,
    dat, obs_Y, file, 
    eval_model_obj=None, 
    preYprob_c=None, preY_c=None,
    plot=True, sheetname='Sheet', print_only=False, notes=None, 
    var_explained_trueZ=None, var_explained_estZ=None, var_explained_V=None, var_explained_VY=None,
    rmse_trueZ=None, rmse_estZ=None, rmse_V=None, rmse_VY=None):

    workbook_name = file

    if ~print_only:
        # Check if the file exists
        if not os.path.isfile(workbook_name):
            # Create a new workbook
            wb = create_modelres_excel_s(file, new_file=True, excel_tab=sheetname)
        else:
            # Load the workbook
            wb = load_workbook(filename=workbook_name)
            # check if the sheet exisits
            if sheetname not in wb.sheetnames:
                # Create a new sheet using the "create_modelres_excel_s" function in order to have the titles for the results
                wb = create_modelres_excel_s(file, new_file=False, excel_tab=sheetname)
                
        page = wb[sheetname]

    # New data to write:
    new_res = [
        datetime.date.today(), datatime_to_string(model.datetime).replace('-', '_'),
        model_type, notes,
        dat['WZ_form'],
        dat['evaluate_on'],
        dat['loss_full'],
        model_train['loss0_train'][-1],
        model_train['loss_train'][-1],
        model_train['loss0_mask1'][-1] if len(model_train['loss0_mask1'])>0 else None, 
        model_train['loss_mask1'][-1] if len(model_train['loss_mask1'])>0 else None, 
        model_train['loss0_mask2'][-1] if len(model_train['loss0_mask2'])>0 else None,  
        model_train['loss_mask2'][-1] if len(model_train['loss_mask2'])>0 else None, 
        # minimal loss on loss0_mask1
        (model_train['loss0_mask1']).min() if len(model_train['loss0_mask1'])>0 else None,
        # the epoch for the minimal loss + how many times of increase before reaching minimum
        model_train['min_loss0_mask1_epoch'], model_train['n_increase_bf_min_loss0_mask1'],
        # loss for WZ:
        model_train['ZWloss_train_all'][-1] if len(model_train['ZWloss_train_all'])>0 else None, 
        model_train['ZWloss_mask1_all'][-1] if len(model_train['ZWloss_mask1_all'])>0 else None, 
        model_train['ZWloss_mask2_all'][-1] if len(model_train['ZWloss_mask2_all'])>0 else None,
        model_train['beta_loss'][-1] if len(model_train['beta_loss'])>0 else None
        ]
    if len(model_train['gamma_loss'])>0:
        new_res = new_res + (model_train['gamma_loss'][-1,]).tolist() 
    else: 
        new_res = new_res + [None, None]  # loss for each gamma 
    # mininum loss for beta and gamma and the correponding epoch
    new_res = new_res + [
        model_train['min_beta_loss'], model_train['min_beta_loss_epoch'], model_train['min_gamma_loss'], model_train['min_gamma_loss_epoch'],
        model_train['min_ZW_loss0_mask1'], model_train['min_ZW_loss0_mask1_epoch']
    ]

    # accuracy + f1 + f1_better
    if eval_model_obj is not None:
        new_res = new_res + [
        eval_model_obj['summary'].loc['train', 'accuracy'], eval_model_obj['summary'].loc['train', 'f1_1'], eval_model_obj['summary'].loc['train', 'f1_1_better'],
        eval_model_obj['summary'].loc['mask1', 'accuracy'], eval_model_obj['summary'].loc['mask1', 'f1_1'], eval_model_obj['summary'].loc['mask1', 'f1_1_better'],
        eval_model_obj['summary'].loc['mask2', 'accuracy'], eval_model_obj['summary'].loc['mask2', 'f1_1'], eval_model_obj['summary'].loc['mask2', 'f1_1_better']
        ]

    new_res = new_res +  [None, None, None]
        
    new_res = new_res + [var_explained_trueZ, var_explained_estZ, var_explained_V, var_explained_VY]
    new_res = new_res + [rmse_trueZ, rmse_estZ, rmse_V, rmse_VY]


    new_res  = new_res + [
        model.Z_orthogonal,
        model_train['min_delta'],
        dat['seed'], dat['mask1_seed'], mask_obj.mask1_percent,
        model.n_gene, model.n_factor, model.px, model.pv, 
        model.W_orthonormal, model.beta_W_orth, model.gamma_W_orth, 
        model.Z_demean_bycenter, model.Z_scale_bycenter,
        model.penalty, model.nquantile,
        model.l1_gamma, model.l1_W, model.l1_beta, 
        model.weight_Y,
        # model_train['lr'], # 
        model_train['optimizer'].param_groups[0]['lr'], 
        model_train["n_epochs"], model_train["early_stop"], model_train["patience"],
        ]
    new_res = new_res + [None, dat['n_sub'], dat['px'], dat['Z_sd']] + dat['center_prop'] + list(dat['gene_prop'].reshape(-1))

    if ~print_only:
        page.append(new_res)
        wb.save(filename=workbook_name)
         
    else:
        print(new_res)


##  check beta estimation accuracy & concordance with the true beta in Simulations
# Y: should be the true observed Y (with NA filled for unobserved entries)
# true_beta: the true beta used in simulating the data (can be the full beta for all Y)
def check_beta_concordance_s(model, center_var, obs_Y, X, true_beta, plot=True):

    centers = np.unique(center_var)
    centers_noMSK = np.delete(centers, np.where(centers == "MSK")) 
    res = pd.DataFrame(columns = centers_noMSK, index = ["discordance", "correlation"])


    # define "real_beta" as the one acturally updated in the model training (correspond to the observed Y) -- code copied from pred_Y(), also update true_beta
    real_beta = copy.deepcopy(model.beta.detach().numpy())
    beta_update = np.empty_like(model.beta.detach().numpy())
    for k in range(X.shape[1]):
        beta_update[:, k] = 1 - np.isnan(np.nanmean(obs_Y.loc[X[:, k]==1, ], 0))
    real_beta[beta_update==0] = np.nan
    # make beta of unobserved genes to nan as well
    true_beta[beta_update==0] = np.nan


    # beta: distribution of beta (blue) and its prediction error (orange)
    if plot:
        plt.hist(true_beta.reshape(-1).reshape(-1), alpha=0.5)
        plt.hist(real_beta.reshape(-1).reshape(-1), alpha=0.5)
        bias_beta = (real_beta - true_beta).reshape(-1)
        plt.hist(bias_beta, alpha=0.5)
        plt.title("True beta (blue), fitted beta (orange) & estimation bias (green)")
        plt.show()
        # display(pd.DataFrame(bias_beta, columns=["bias_beta"]).describe())
    # rMSE
    rmse = np.sqrt(mean_squared_error(real_beta[~np.isnan(real_beta)], true_beta[~np.isnan(true_beta)]))
    print("** rMSE of fitted beta", rmse)
    
    # beta: sign same
    sign_same = np.sum((true_beta>0) == (real_beta>0)) / np.prod(real_beta.shape)
    print("** Beta sign same (%):", sign_same)

    # Beta difference across center consistency:
    beta_center_concordance = np.unique((real_beta[:,0] - real_beta[:,1] > 0) == (true_beta[:,0] - true_beta[:,1] > 0), return_counts = True) # note: NA will be counted as concordance in this case
    beta_center_concordance = 1 - beta_center_concordance[1][0] / (beta_update.sum(1)==2).sum()  # 1 - discordance among shared genes
    print("** beta_center_concordance (%):", beta_center_concordance)
 
    return {'real_beta': real_beta, 'true_beta': true_beta, 'rmse':rmse, 'sign_same':sign_same, 'center_concordance':beta_center_concordance}







